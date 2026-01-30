"""
PDF Export Module for BDX Analysis

This module provides classes and functions for exporting ROOT histograms to PDF files
with organized naming schemes and directory structures.
"""

import ctypes
import glob
import re
from pathlib import Path
from typing import Optional, Tuple, List
import csv
import ROOT
from utils import setup_root
from simulation_summary import SimulationSummary


class Exporter:
    """
    Class to handle export of ROOT histograms with organized file structure.
    """
    
    def __init__(self, style_env: Optional[Tuple] = None, verbose: bool = True, save_macro: bool = False):
        """
        Initialize the exporter.
        
        Args:
            style_env: Style environment tuple (set_style, style_h1, style_h2, quiet) or None
            verbose: Print integral information for 1D 'lin' histograms
            save_macro: Also save each canvas as a ROOT macro (.C)
        """
        self.verbose = verbose
        self.style_env = style_env
        self.save_macro = save_macro
        
        # Setup style environment if not provided
        if self.style_env is None:
            set_style, stileh1, stileh2, quiet = setup_root()
            self.style_env = (set_style, stileh1, stileh2, quiet)
    
    def export_single_file(self, root_file_path: str, output_dir: str = "./Analysis") -> None:
        """
        Export all histograms from a single ROOT file to PDF files.
        
        Args:
            root_file_path: Path to the ROOT file
            output_dir: Output directory for PDF files
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Reuse unified exporting implementation
        self._export_histograms_from_file(Path(root_file_path), output_dir=output_dir, name_prefix="", indent_print="")
    
    def export_all_analysis_files(self, output_dir: str, particle_name: Optional[str] = None, 
                                 energy_ranges: Optional[List[Tuple[float, float]]] = None) -> None:
        """
        Export all histograms from multiple analysis ROOT files to organized PDF files.
        
        Args:
            output_dir: Output directory containing the ROOT files
            particle_name: Particle name to filter ROOT files (optional)
            energy_ranges: List of (min_energy, max_energy) tuples in GeV for proper ordering
        """
        output_dir = Path(output_dir)
        
        # Create plots subdirectory
        plots_dir = output_dir / "plots"
        plots_dir.mkdir(parents=True, exist_ok=True)
        
        # Find and sort all analysis ROOT files (exclude simulation_summary.root)
        root_files = self._get_sorted_analysis_files(output_dir, particle_name, energy_ranges)
        if not root_files:
            print(f"No analysis ROOT files found in {output_dir}")
            return
        
        print(f"Found {len(root_files)} analysis ROOT files")
        print(f"Exporting histograms to PDF in {plots_dir}")
        
        # Process each ROOT file
        for i, root_file_path in enumerate(root_files):
            print(f"\nProcessing file {i+1}/{len(root_files)}: {root_file_path.name}")
            
            # Extract binning and energy range from filename
            binning_energy_info = self._extract_binning_energy_from_filename(root_file_path.stem)
            
            # Create prefix for ordering (00, 01, 02, etc.) + binning/energy info
            prefix = f"{i:02d}_{binning_energy_info}_"
            
            # Process the ROOT file
            self._export_histograms_from_file(root_file_path, output_dir=str(plots_dir), name_prefix=prefix, indent_print="    ")
        
        print(f"\nCompleted PDF export. All plots saved in: {plots_dir}")

    def export_histogram_statistics(self, output_dir: str,
                                    particle_name: Optional[str] = None,
                                    energy_ranges: Optional[List[Tuple[float, float]]] = None,
                                    excel_filename: str = "histogram_statistics.xlsx",
                                    surface_dims: Optional[Tuple[float, float, float, float]] = None) -> None:
        """
        Export histogram statistics to an Excel file (CSV fallback if Excel engine not available).

        The statistics are computed for 1D histograms whose name contains 'lin' to match the existing console statistics behavior.

        Columns: analyzed_file, detector, histogram_name, energy_range, integral, error, error_percent,
                 xl, xh, yl, yh, delta_x, delta_y

        Args:
            output_dir: Output directory containing the ROOT files
            particle_name: Particle name to filter ROOT files (optional)
            energy_ranges: List of (min_energy, max_energy) tuples in GeV for ordering and labeling
            excel_filename: Output Excel filename (created under output_dir)
        """
        from utils import format_energy

        base_dir = Path(output_dir)

        # Identify and sort ROOT files using the shared helper
        sorted_files = self._get_sorted_analysis_files(base_dir, particle_name, energy_ranges)

        if not sorted_files:
            print(f"No analysis ROOT files found in {base_dir}")
            return

        rows = []
        energy_rows = {}
        energy_order = []

        def extract_detector_from_histogram_name(name: str) -> str:
            """
            Extract detector identifier from histogram names such as:
            - neutrons_h1_fl_1000_enclose_lin
            - neutrons_h2_fl_1000_enclose
            """
            try:
                after = name
                for marker in ("h1_", "h2_"):
                    if marker in after:
                        after = after.split(marker, 1)[1]
                        break

                for suffix in ("_lin", "_log"):
                    if after.endswith(suffix):
                        after = after[: -len(suffix)]
                        break
                return after
            except Exception:
                return "unknown"

        # Define integration parameters used below so we can report them in the Notes sheet
        integration_first_bin = 1
        integration_option = "width"  # Passed to IntegralAndError's option argument
        # Try to read EOT from simulation summary file in the output directory
        eot_value = None
        try:
            summary_path = base_dir / "simulation_summary.root"
            if summary_path.exists():
                eot_value = SimulationSummary.read_eot_from_file(str(summary_path))
        except Exception:
            eot_value = None

        for file_path in sorted_files:
            # Derive energy range label using shared helper
            energy_label = self._derive_energy_label(file_path, energy_ranges)

            root_file = ROOT.TFile(str(file_path), "READ")
            if not root_file or root_file.IsZombie():
                print(f"Warning: Could not open ROOT file: {file_path}")
                continue

            detector_rows = {}
            keys = root_file.GetListOfKeys()
            for key in keys:
                obj = key.ReadObj()
                if not obj.InheritsFrom("TH1"):
                    continue

                hist_name = obj.GetName()
                detector_name = extract_detector_from_histogram_name(hist_name)
                is_2d = obj.InheritsFrom("TH2")

                if is_2d:
                    error = ctypes.c_double(0)
                    last_bin_x = obj.GetNbinsX()
                    last_bin_y = obj.GetNbinsY()
                    integral = obj.IntegralAndError(
                        integration_first_bin,
                        last_bin_x,
                        integration_first_bin,
                        last_bin_y,
                        error,
                        integration_option,
                    )
                    percent_error = (error.value / integral) * 100 if integral != 0 else 0.0

                    x_axis = obj.GetXaxis()
                    y_axis = obj.GetYaxis()
                    xl = x_axis.GetXmin()
                    xh = x_axis.GetXmax()
                    yl = y_axis.GetXmin()
                    yh = y_axis.GetXmax()

                    detector_rows[detector_name] = {
                        "priority": 0,
                        "data": {
                            "analyzed_file": file_path.name,
                            "detector": detector_name,
                            "histogram_name": hist_name,
                            "energy_range": energy_label,
                            "integral": float(integral),
                            "error": float(error.value),
                            "error_percent": float(percent_error),
                            "xl": float(xl),
                            "xh": float(xh),
                            "yl": float(yl),
                            "yh": float(yh),
                            "delta_x": float(xh - xl),
                            "delta_y": float(yh - yl),
                        },
                    }
                    continue

                if "lin" not in hist_name:
                    continue

                error = ctypes.c_double(0)
                last_bin = obj.GetNbinsX()
                integral = obj.IntegralAndError(
                    integration_first_bin,
                    last_bin,
                    error,
                    integration_option,
                )
                percent_error = (error.value / integral) * 100 if integral != 0 else 0.0

                x_axis = obj.GetXaxis()
                xl = x_axis.GetXmin()
                xh = x_axis.GetXmax()
                yl = yh = None
                delta_x = xh - xl
                delta_y = None

                candidate = {
                    "priority": 1,
                    "data": {
                        "analyzed_file": file_path.name,
                        "detector": detector_name,
                        "histogram_name": hist_name,
                        "energy_range": energy_label,
                        "integral": float(integral),
                        "error": float(error.value),
                        "error_percent": float(percent_error),
                        "xl": float(xl) if xl is not None else None,
                        "xh": float(xh) if xh is not None else None,
                        "yl": float(yl) if yl is not None else None,
                        "yh": float(yh) if yh is not None else None,
                        "delta_x": float(delta_x) if delta_x is not None else None,
                        "delta_y": float(delta_y) if delta_y is not None else None,
                    },
                }

                existing = detector_rows.get(detector_name)
                if existing is None or candidate["priority"] < existing["priority"]:
                    detector_rows[detector_name] = candidate

            for entry in detector_rows.values():
                row = entry["data"]
                rows.append(row)
                energy_label_key = row["energy_range"]
                if energy_label_key not in energy_rows:
                    energy_rows[energy_label_key] = []
                    energy_order.append(energy_label_key)
                energy_rows[energy_label_key].append(row)

            root_file.Close()

        if not rows:
            print("No matching histograms found for statistics export.")
            return

        # Try Excel via pandas/openpyxl; otherwise fallback to CSV
        excel_path = base_dir / excel_filename
        try:
            import pandas as pd  # type: ignore
            columns = [
                "analyzed_file", "detector", "histogram_name", "energy_range",
                "integral", "error", "error_percent",
                "xl", "xh", "yl", "yh", "delta_x", "delta_y"
            ]
            with pd.ExcelWriter(excel_path) as writer:
                for idx, label in enumerate(energy_order):
                    df = pd.DataFrame(energy_rows[label], columns=columns)
                    sheet_name = f"{idx:02d}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                # Add Notes sheet with integral description (dynamic by integration params)
                if eot_value is not None:
                    eot_note = f"scaled by 1/EOT where EOT={eot_value}"
                else:
                    eot_note = "scaled by 1/EOT"
                notes_text = (
                    f"Integrals computed with IntegralAndError(first_bin={integration_first_bin}, "
                    f"last_bin=N_bins_x, option=\"{integration_option}\"), {eot_note}."
                )
                pd.DataFrame({"Notes": [notes_text]}).to_excel(writer, sheet_name="Notes", index=False, header=False)

            # Apply number formats using openpyxl if available, to each sheet
            try:
                import openpyxl  # type: ignore
                wb = openpyxl.load_workbook(excel_path)
                sci_format = '0.00E+00'
                percent_literal_format = '0.00"%"'  # Do not scale values, only show % sign
                for ws in wb.worksheets:
                    if ws.title == "Notes":
                        continue
                    header_to_col = {}
                    for col in range(1, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if isinstance(header, str):
                            header_to_col[header] = col
                    for row_idx in range(2, ws.max_row + 1):
                        if "integral" in header_to_col:
                            ws.cell(row=row_idx, column=header_to_col["integral"]).number_format = sci_format
                        if "error" in header_to_col:
                            ws.cell(row=row_idx, column=header_to_col["error"]).number_format = sci_format
                        if "error_percent" in header_to_col:
                            ws.cell(row=row_idx, column=header_to_col["error_percent"]).number_format = percent_literal_format
                        for dim_col in ["xl", "xh", "yl", "yh", "delta_x", "delta_y"]:
                            if dim_col in header_to_col:
                                ws.cell(row=row_idx, column=header_to_col[dim_col]).number_format = '0.00'
                # Move Notes sheet to the first position
                if "Notes" in wb.sheetnames:
                    notes_ws = wb["Notes"]
                    wb._sheets.remove(notes_ws)
                    wb._sheets.insert(0, notes_ws)
                wb.save(excel_path)
            except Exception:
                # If formatting fails, leave the file as written
                pass

            print(f"Histogram statistics saved to: {excel_path}")
        except Exception as exc:
            # Fallback to CSV
            csv_path = excel_path.with_suffix('.csv')
            with open(csv_path, mode='w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=[
                    "analyzed_file", "detector", "histogram_name", "energy_range",
                    "integral", "error", "error_percent",
                    "xl", "xh", "yl", "yh", "delta_x", "delta_y"
                ])
                writer.writeheader()
                writer.writerows(rows)
            print(f"Excel export unavailable ({exc}). CSV written to: {csv_path}")
    
    def _sort_files_by_energy_ranges(self, root_files: List[Path], energy_ranges: List[Tuple[float, float]]) -> List[Path]:
        """
        Sort ROOT files by energy ranges using configuration data.
        
        The full range (largest span) goes first as "00_",
        then sub-ranges in increasing order of minimum energy.
        
        Args:
            root_files: List of ROOT file paths
            energy_ranges: List of (min_energy, max_energy) tuples in GeV
            
        Returns:
            Sorted list of ROOT file paths
        """
        # Create mapping from energy range to file
        file_energy_map = {}
        
        # Calculate energy spans for ordering
        energy_spans = [(max_e - min_e, min_e, max_e) for min_e, max_e in energy_ranges]
        
        # Find the largest span (full range)
        max_span = max(energy_spans, key=lambda x: x[0])[0]
        
        # Sort energy ranges: full range first (largest span), then sub-ranges by min energy (ascending)
        def sort_key(x):
            span, min_e, max_e = x
            if span == max_span:
                return (0, 0)  # Full range goes first
            else:
                return (1, min_e)  # Sub-ranges sorted by minimum energy (ascending)
        
        sorted_ranges = sorted(energy_spans, key=sort_key)
        
        # For each sorted energy range, find the corresponding file
        sorted_files = []
        for span, min_e, max_e in sorted_ranges:
            # Find the file that matches this energy range
            for file_path in root_files:
                # Extract energy info from filename to match with config
                if self._file_matches_energy_range(file_path, min_e, max_e):
                    if file_path not in sorted_files:  # Avoid duplicates
                        sorted_files.append(file_path)
                    break
        
        # Add any remaining files that weren't matched
        for file_path in root_files:
            if file_path not in sorted_files:
                sorted_files.append(file_path)
        
        return sorted_files
    
    def _file_matches_energy_range(self, file_path: Path, min_energy: float, max_energy: float) -> bool:
        """
        Check if a ROOT file matches the given energy range by examining the filename.
        
        Args:
            file_path: Path to the ROOT file
            min_energy: Minimum energy in GeV
            max_energy: Maximum energy in GeV
            
        Returns:
            True if the filename matches the energy range
        """
        from utils import format_energy
        
        # Convert energies to formatted strings that match the filename
        min_str = format_energy(min_energy)
        max_str = format_energy(max_energy)
        
        # Check if both energy strings appear in the filename
        filename = file_path.name
        return min_str in filename and max_str in filename
    
    def _extract_binning_energy_from_filename(self, file_stem: str) -> str:
        """
        Extract binning and energy range information from ROOT filename.
        
        Example: analysis_neutrons_1000bins_0.1eV_6.0GeV.root
        Returns: 1000bins_0.1eV_6.0GeV
        """
        # Use regex to extract the pattern: {N}bins_{min_energy}_{max_energy}
        pattern = r'(\d+bins_[\d.]+\w+_[\d.]+\w+)'
        match = re.search(pattern, file_stem)
        
        if match:
            return match.group(1)
        else:
            # Fallback: try to extract just the part after the particle name
            parts = file_stem.split('_')
            if len(parts) >= 4:
                # Expect format: analysis_particlename_binning_minE_maxE
                return '_'.join(parts[2:])  # Return binning_minE_maxE
            else:
                return "unknown"
    
    def _export_file_with_prefix(self, root_file_path: Path, plots_dir: Path, prefix: str) -> None:
        """
        Helper method to export histograms from a single ROOT file with organized naming.
        """
        # Delegate to unified implementation; keep method for backward-compatibility inside class
        self._export_histograms_from_file(root_file_path, output_dir=str(plots_dir), name_prefix=prefix, indent_print="    ")
    
    def _print_histogram_statistics(self, histogram, hist_name: str, indent: str = "") -> None:
        """
        Print statistical information for a histogram.
        
        Args:
            histogram: ROOT histogram object
            hist_name: Name of the histogram
            indent: Indentation string for output formatting
        """
        error = ctypes.c_double(0)
        if histogram.InheritsFrom("TH2"):
            integral = histogram.IntegralAndError(
                1,
                histogram.GetNbinsX(),
                1,
                histogram.GetNbinsY(),
                error,
                "width",
            )
        else:
            integral = histogram.IntegralAndError(
                1,
                histogram.GetNbinsX(),
                error,
                "width",
            )
        percent_error = (error.value / integral) * 100 if integral != 0 else 0
        entries = histogram.GetEntries()
        
        print(f"{indent}{hist_name}: {integral:.2E} ± {error.value:.2E} ({percent_error:.2f}%)")
        print(f"{indent}  Entries: {entries}")

    # -----------------------
    # Internal helper methods
    # -----------------------

    def _get_sorted_analysis_files(self, base_dir: Path, particle_name: Optional[str], energy_ranges: Optional[List[Tuple[float, float]]]) -> List[Path]:
        """
        List and sort analysis ROOT files under base_dir, optionally filtered by particle name.
        Excludes any file containing 'simulation_summary' in its name.
        Sorts by provided energy ranges if available, otherwise by filename.
        """
        if particle_name:
            pattern = f"analysis_{particle_name}_*.root"
        else:
            pattern = "analysis_*.root"

        root_files = list(Path(base_dir).glob(pattern))
        root_files = [f for f in root_files if "simulation_summary" not in f.name]

        if not root_files:
            return []

        if energy_ranges:
            return self._sort_files_by_energy_ranges(root_files, energy_ranges)
        return sorted(root_files, key=lambda f: f.name)

    def _derive_energy_label(self, file_path: Path, energy_ranges: Optional[List[Tuple[float, float]]]) -> str:
        """
        Derive a human-readable energy range label for a given analysis file using
        configured ranges when available, otherwise fall back to parsing the filename.
        """
        from utils import format_energy

        if energy_ranges:
            for min_e, max_e in energy_ranges:
                if self._file_matches_energy_range(file_path, min_e, max_e):
                    return f"{format_energy(min_e)} - {format_energy(max_e)}"
            return "unknown"

        info = self._extract_binning_energy_from_filename(file_path.stem)
        parts = info.split("_")
        if len(parts) >= 3:
            return f"{parts[-2]} - {parts[-1]}"
        return "unknown"

    def _export_histograms_from_file(self, root_file_path: Path, output_dir: str, name_prefix: str, indent_print: str) -> None:
        """
        Unified implementation to open a ROOT file, iterate histograms, render and save PDFs,
        and optionally print statistics for 1D 'lin' histograms.
        """
        destination_dir = Path(output_dir)
        destination_dir.mkdir(parents=True, exist_ok=True)

        root_file = ROOT.TFile(str(root_file_path), "READ")
        if not root_file or root_file.IsZombie():
            warning_indent = indent_print if indent_print else ""
            print(f"{warning_indent}Warning: Could not open ROOT file: {root_file_path}")
            return

        keys = root_file.GetListOfKeys()
        set_style, style_h1, style_h2, quiet = self.style_env

        histogram_count = 0
        for key in keys:
            obj = key.ReadObj()
            if not obj.InheritsFrom("TH1"):
                continue

            is_2d = obj.InheritsFrom("TH2")
            hist_name = obj.GetName()
            pdf_filename = f"{name_prefix}{hist_name}.pdf"
            pdf_path = destination_dir / pdf_filename

            style = style_h2 if is_2d else style_h1
            with set_style(style):
                canvas = ROOT.TCanvas(hist_name, hist_name, 700, 600)
                if is_2d:
                    canvas.SetLogz()
                    obj.Draw("colz")
                else:
                    canvas.SetLogy()
                    # If this histogram corresponds to the "log" variant from config, set X axis to log scale
                    try:
                        if ("_log" in hist_name) and ("_lin" not in hist_name):
                            canvas.SetLogx()
                    except Exception:
                        pass
                    obj.Draw("hist")
                with quiet:
                    canvas.SaveAs(str(pdf_path))
                    if self.save_macro:
                        macro_path = pdf_path.with_suffix('.C')
                        canvas.SaveAs(str(macro_path))
                with quiet:
                    canvas.SaveAs(str(pdf_path))
                    if self.save_macro:
                        macro_path = pdf_path.with_suffix('.C')
                        canvas.SaveAs(str(macro_path))

            histogram_count += 1

            # Print integral for 1D 'lin' histograms
            if (not is_2d) and ("lin" in hist_name) and self.verbose:
                self._print_histogram_statistics(obj, hist_name, indent=indent_print)

        if indent_print:
            print(f"{indent_print}Exported {histogram_count} histograms from {root_file_path.name}")
        root_file.Close()


# Convenience functions for backward compatibility and easy imports
def export_histograms_to_pdf(root_file_path: str, 
                            output_dir: str = "./Analysis",
                            style_env: Optional[Tuple] = None,
                            verbose: bool = True) -> None:
    """
    Export all histograms from a single ROOT file to PDF files.
    
    Args:
        root_file_path: Path to the ROOT file
        output_dir: Output directory for PDF files
        style_env: Style environment tuple or None
        verbose: Print integral information for 1D 'lin' histograms
    """
    exporter = Exporter(style_env=style_env, verbose=verbose)
    exporter.export_single_file(root_file_path, output_dir)


def export_all_analysis_histograms_to_pdf(output_dir: str,
                                         particle_name: Optional[str] = None,
                                         style_env: Optional[Tuple] = None,
                                         verbose: bool = True,
                                         energy_ranges: Optional[List[Tuple[float, float]]] = None,
                                         save_macro: bool = False) -> None:
    """
    Export all histograms from multiple analysis ROOT files to organized PDF files.
    
    Args:
        output_dir: Output directory containing the ROOT files
        particle_name: Particle name to filter ROOT files (optional)
        style_env: Style environment tuple or None
        verbose: Print integral information for 1D 'lin' histograms
        energy_ranges: List of (min_energy, max_energy) tuples in GeV for proper ordering
    """
    exporter = Exporter(style_env=style_env, verbose=verbose, save_macro=save_macro)
    exporter.export_all_analysis_files(output_dir, particle_name, energy_ranges)


def export_histogram_statistics_to_excel(output_dir: str,
                                         particle_name: Optional[str] = None,
                                         style_env: Optional[Tuple] = None,
                                         verbose: bool = True,
                                         energy_ranges: Optional[List[Tuple[float, float]]] = None,
                                         excel_filename: str = "histogram_statistics.xlsx",
                                         surface_dims: Optional[Tuple[float, float, float, float]] = None) -> None:
    """
    Export histogram statistics to an Excel file (CSV fallback if Excel engine not available).

    Columns include analyzed_file, detector, histogram_name, energy_range, integral, error, error_percent,
    and optionally xl, xh, yl, yh, delta_x, delta_y when surface_dims is provided.
    """
    exporter = Exporter(style_env=style_env, verbose=verbose)
    exporter.export_histogram_statistics(output_dir, particle_name, energy_ranges, excel_filename, surface_dims)
